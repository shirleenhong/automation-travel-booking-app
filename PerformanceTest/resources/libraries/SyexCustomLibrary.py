# -*- coding: utf-8 -*-
from collections import Counter, OrderedDict
from robot.libraries.String import String
from robot.libraries.BuiltIn import BuiltIn
from robot.api import logger
from Constant import CurrencyConstant
from Constant import CountryConstant
from SyexUiaLibrary import SyexUiaLibrary
from SyexDateTimeLibrary import SyexDateTimeLibrary
from RailLibrary import RailLibrary
from pywinauto.application import Application
from pywinauto.controls.win32_controls import ComboBoxWrapper
from pywinauto.findbestmatch import MatchError
from uialibrary import Win32API as win32api
import os
import sys
import getpass
import math 
import re
import win32clipboard
import autoit
import time
import pymssql
import uialibrary
from openpyxl import load_workbook
from decimal import Decimal, ROUND_HALF_UP, ROUND_HALF_DOWN, ROUND_DOWN, ROUND_UP
reload(sys)
sys.setdefaultencoding('Cp1252')

class SyexCustomLibrary:

    ROBOT_LIBRARY_SCOPE = 'Global'

    def __init__(self):
        self.app = Application()
        self.builtin = BuiltIn()

    def _create_app_instance(self):
        appconnect = self.app.Connect(title_re='Power Express')        
        # appconnect = self.app.Connect(path='PowerExpress.exe')
        app = appconnect[u'WindowsForms10.Window.8.app.0.3309ded_r17_ad1']
        app.Wait('ready')
        return app

    def launch_power_express(self, version, syex_env, username, use_mock_env="False"):
        if use_mock_env == "True":
            use_mock_env = " MockGdsLookup MockPortrait"
        else:
            use_mock_env = ""

        express_path = self.get_power_express_path(version)
        express_exec = 'PowerExpress.exe' + ' ENV:' + syex_env +' testuser:' + username + use_mock_env +''
        os.chdir(express_path)
        try:
            autoit.run(express_exec)
        except Exception as e:
            self.builtin.fail(str(e) + '. Check if PowerExpress is installed correctly')         
        autoit.process_wait('PowerExpress.exe')
        self.builtin.set_test_variable('${username}', username)
        self.builtin.set_suite_variable('${syex_env}', syex_env)

    def get_power_express_path(self, version):
        express_basedir = "C:\\Program Files (x86)\\Carlson Wagonlit Travel\\Power Express "
        express_path = "{}{}".format(express_basedir, version)
        if os.path.exists(express_path):
            return express_path
        else:
            self.builtin.fail('Path not found. Change your Express installation to default location')

    def get_log_path(self, test_environment, version):
        local_log_path = os.getenv('LOCALAPPDATA') + "\\Carlson Wagonlit Travel\\Power Express\\"+ version +"\\"
        citrix_log_path = "D:\\Syex_Logs\\"
        if test_environment == 'citrix':
            return citrix_log_path
        else:
            return local_log_path

    def get_home_dir_path(self):
        return os.path.expanduser('~')

    def get_UserName(self):
        return self.get_local_username()

    def get_local_username(self):
        return getpass.getuser()

    def select_profile(self, user_profile):
        try:
            select_profile_app = self.app.Connect(title_re='Power Express')
            autoit.win_wait('[REGEXPTITLE:rofil]', 60)
            autoit.win_activate('[REGEXPTITLE:rofil]')
            select_profile_window = select_profile_app.Window_(title_re=u'.*rofil')
            select_profile_window.Wait('ready')
            select_profile_window.SetFocus()
            select_profile_window_list_box = select_profile_window.ListBox
            select_profile_window_list_box.Select(user_profile)
            select_profile_ok_button = select_profile_window.OK
            select_profile_ok_button.Click()
            self.builtin.set_suite_variable('${user_profile}', user_profile)
        except Exception:
            pass
        time.sleep(2)
        try:
            if autoit.control_command('Power Express', '[NAME:YesBtn]', 'IsVisible'):
                print('ok')
                autoit.control_click('Power Express', '[NAME:YesBtn]')
        except Exception:
            pass

    def select_team_name(self, team_name):
        SyexUiaLibrary().toggle_checkbox(team_name)

    def get_team_index_value(self, team_name, user_profile, syex_env='Test', window_title=None):
        # if syex_env.capitalize() != 'Test':
        #    team_list = [x.Name for x in uialibrary.ListControl(AutomationId='chklistTeam').GetChildren()]
        #     if 'Vertical' in team_list[0]:
        #         team_list.remove(team_list[0])
        # else:
        #     conn = pymssql.connect('VWNV02DS01463.int.carlsonwagonlit.com', 'SYMPHONIEEXPRESS', 'SyExAugust17*', "Desktop_Test")
        #     cursor = conn.cursor()
        #     declaration = "DECLARE @UprofileGUID as Varchar(10)= (Select SystemUserGuid from Systemuser where UserProfileIdentifier='{}')"\
                    # .format(user_profile)
        #     stored_proc = "{}\nEXEC dbo.spExpress_GetTeamList_v3 @SystemUserGuid=@UprofileGUID".format(declaration)
        #     cursor.execute(stored_proc)
        #     team_list = [r[1] for r in cursor]
        #     print team_list
        #     conn.close()
        #     try:
        #         team_name_index = team_list.index(team_name)
        #     except Exception:
        #         self.builtin.fail("Team '{}' is not found on team list".format(team_name))
        team_list = [x.Name for x in uialibrary.ListControl(AutomationId='chklistTeam')\
                    .GetChildren() if x.Name != "Vertical"]
        team_name_index = team_list.index(team_name)
        logger.info("Team index is '{}'".format(team_name_index))
        return team_name_index
       
    def get_all_panels(self):
        """ 
        Returns all active / current panels

        Example:
        | @{current_panels} = | Get All Panels |

        =>
        @{current_panels} =  [ Client Info | Client Fees ]
        """
        syexuia = SyexUiaLibrary().create_power_express_handle()
        return [item.Name for item in uialibrary.TabControl(searchFromControl=syexuia).GetChildren()]

    def select_value_from_dropdown_list(self, control_id, value_to_select, window_title='Power Express', by_index=False):
        """ 
        Selects dropdown value using text or index.  If using index, use by_index=True
        
        Example:
        | Select Value From Dropdown List | [NAME:ccboClass_1] | Class Code Sample |
        """
        logger.info("Selecting dropdown value '{}'".format(value_to_select))
        if str(by_index) == "False":
            syexuia = SyexUiaLibrary()
            syexuia.select_dropdown_value(control_id, value_to_select)
        else:
            try:
                autoit.control_click(window_title, control_id)
                autoit.send("{PGUP}")
                autoit.send('{BACKSPACE}')
                autoit.send('{HOME}')
            except Exception:
                self.builtin.fail("Control '{}' not found".format(control_id))
            autoit.control_focus(window_title, control_id)
            if value_to_select == "0":
                autoit.send('{ENTER}')
                autoit.send('{TAB}')
            else:
                autoit.send("{DOWN}" * int(value_to_select))
                autoit.send('{ENTER}')
                autoit.send('{TAB}')

    def get_value_from_dropdown_list(self, control_id, window_title='Power Express'):
        """ 
        Returns a list containing all values of the specified dropdown.
        The default dropdown value is retained.
        """
        syexuia = SyexUiaLibrary() 
        return syexuia.get_dropdown_values(control_id)

    def select_value_from_combobox(self, combobox_field, combobox_value):
        """ 
        Description:
        Selects value from dropdown using dropdown name or control id as argument
        Usage:
        | Select Value From Combobox | Bundled Fee | Apply Bundled Fee |
        """
        if bool(re.search(r"(\[.*:)|(\])", combobox_field)):
            uia = SyexUiaLibrary()
            uia.select_dropdown_value(combobox_field, combobox_value)
        else:
            self._select_combobox_using_combobox_name(combobox_field, combobox_value)

    def _select_combobox_using_combobox_name(self, combobox_name, combobox_value):
        logger.info("Selecting '{}' from '{}' combobox".format(combobox_value, combobox_name))            
        try:
            self._create_app_instance()[u''.join(combobox_name)+'ComboBox'].Select(u''.join(combobox_value))
        except (ValueError):
            self.builtin.fail("Combobox value '{}' is not found".format(combobox_value))

    def get_value_from_combobox(self, combobox_name):
        """ 
        Description:
        Gets all dropdown values/items using dropdown name as argument

        Usage:
        | Get Value From Combobox | Bundled Fee |
        
        =>
        [u'Apply Bundled Fee', u'No Bundled Fee']
        """
        try:
            return self._create_app_instance()[u''.join(combobox_name)+'ComboBox'].ItemTexts()
        except Exception:
            self.builtin.fail("Usage of keyword is incorrect, try to use other keyword")

    def select_value_from_listbox(self, listbox_value):
        uia = SyexUiaLibrary().connect_to_express_via_uia()
        listbox = uia.ListBox
        listbox.Select(u''.join(listbox_value))

    def select_tst_fare(self, fare_number):
        """ 
        Description:
        For Rail Panel
        Select Tab using given EXACT value
        """
        listbox = self._create_app_instance().ListBox
        listbox.Select(u'  %s' % fare_number)

    def get_currency(self, country_code):
        return getattr(CurrencyConstant, country_code.upper())

    def get_percentage_value(self, number):
        percentage = (float(number) / float(100))
        return percentage

    def convert_to_float(self, value, precision=2):
        """ 
        Converts value to float. By default precision used is 2.

        Example:
        | ${convered_value} = | Convert To Float | 15.0909 | precision=5 |

        """
        try:
            return "{:.{prec}f}".format(float(value), prec=precision)
        except Exception as error:
            raise ValueError(repr(error))

    def get_visible_tab(self):
        """ 
        Returns fare tabs in list data type

        """
        tab_list = [tab.Name for tab in uialibrary.TabControl(AutomationId='TabControl1').GetChildren() 
                    if tab.ControlTypeName == 'TabItemControl']
        return tab_list

    def get_string_matching_regexp(self, reg_exp, data, group_number=0):
        """ 
        Get string matching Reg Expression.

        Example:
        | ${string} = | Get String Matching Regexp | TAX [0-9]+\.[0-9][0-9] | NZD120.00 TAX12.00 |
        
        Output:
        ${string} = TAX12.00
        
        """
        try:
            reg_exp = re.compile(reg_exp)
            m = re.search(reg_exp, data)
            return m.group(int(group_number))
        except Exception:
            return 0

    def get_string_using_marker(self, whole_string, first_marker, end_marker=None):
        """ 
        Description:
        Returns substring or in between string of first marker and end marker.
        Leading and trailing spaces will be stripped.

        Usage:
        | ${summary_texts} = | Get String Using Marker | HP*BIRTHDATE-20JUN84 98| - | ${SPACE} |

        =>
        ${summary_texts} = 20JUN84
        """
        self.builtin.log("Getting string from {} between {} and {}".format(whole_string, first_marker, end_marker))
        start = str(whole_string).find(first_marker) + len(first_marker)
        end = whole_string.find(end_marker, start)
        if end == -1:
            end = len(whole_string)
            return whole_string[start:end].strip()
        else:
            return whole_string[start:end].strip()      

    def get_minimum_value_from_list(self, list_):
        """ 
        Description:
        Returns the least / minimum value of list item

        Usage:
        | @{fare_list}          = | Create List                 | 12.3 | 21.2 | 35.3 | 10.8 |
        | ${minimum_fare_value} = | Get Minimum Value From List | ${fare_list}              | 

        Result:
        | ${minimum_fare_value} = | 10.8 |

        """
        try:
            return min(float(value) for value in list_)
        except ValueError:
            self.builtin.fail("List does not contain any value")

    def get_required_flight_details(self, raw_flight_details, city_pair_marker):
        raw_flight_details = raw_flight_details.strip()
        flight_details_line = raw_flight_details.find(city_pair_marker)
        fare_line_length = flight_details_line + len(city_pair_marker)
        raw_flight_details = raw_flight_details[:fare_line_length]
        flight_details_list = raw_flight_details.split(" ")
        flight_details_list = [x for x in flight_details_list if x]
        del flight_details_list[0]
        return flight_details_list

    def round_up_hk(self, number):
        """
        Description:
        Returns round up value in integer if decimal is found otherwise no roundup 
        """
        number_decimal = str(number-int(number))[2:]
        if float(number_decimal) == 0:
            return int(number)
        else:
            return int(number) + 1

    def round_off(self, number, precision=2):
        """
        Description:
        Returns round off to 2 decimal places. By default the precision is two (2)

        Usage:
        | ${number} | Round Off | 12.8998 |

        Output:
        ${number} = 12.90
        """
        try:
            number_decimal = round(float(number), int(precision))
            return str(number_decimal)
        except (ValueError):
            self.builtin.fail("'{}' is an invalid input.".format(number))

    def click_add_manual_fare(self):
        uialibrary.ButtonControl(AutomationId='btnAddManualFare').Click()

    def click_add_alternate_fare(self):
        uialibrary.ButtonControl(AutomationId='btnAddAlternateFare').Click()        

    def click_remove_fare(self):
        uialibrary.ButtonControl(AutomationId='btnRemoveFare').Click()           

    def get_tst_list(self):
        """ 
        Description:
        Get all TST item and save it to list type

        Usage:
        | @{tst_list} = | get_tst_list |

        Result:
        | @{list} =     | 1 | 2 |
        """    
        tst_listbox = self._create_app_instance().ListBox
        tst_list =  tst_listbox.ItemTexts()
        tst_list = [item.strip() for item in tst_list]
        return tst_list

    def remove_duplicate_from_list(self, list_):
        """ 
        Description:
        Remove all duplicate item from list

        Usage:
        | @{sample_list} = | Create List                | 1  | 2 | 3 | 1 |
        | @{sample_list} = | Remove Duplicate From List | ${sample_list} | 

        Result:
        | @{sample_list} = | [ 1, 2, 3 ] |
        """   
        seen = set()
        seen_add = seen.add
        return [value for value in list_ if not (value in seen or seen_add(value))]

    def replace_value_in_list(self, list_, index_or_value, new_value):
        if index_or_value.isdigit():
            try:
                list_[int(index_or_value)] = new_value
            except IndexError:
                self.builtin.fail("Index value {} is incorrect".format(index_or_value))
        else:
            try:
                index = list_.index(index_or_value)
                list_[index] = new_value
            except ValueError:
                self.builtin.fail("{} is not in list: Actual list values are : {}".\
                    format(index_or_value, ", ".join(list_)))
        return list_

    def click_menu_item(self, menu_title, double_click=False, window="Power Express"):
        """ 
        Description:
        Selects Menu item such as "Change Team", "Manual Contact", etc.
        
        Usage:
        | Click Menu Item | Change Team |

        """
        # app_dlg = self._create_power_express_backend_app()
        uia = SyexUiaLibrary().connect_to_express_via_uia()        
        menu_item = uia.child_window(title=menu_title, control_type="Button")
        if double_click == "True":
            menu_item.click_input(double=True)
        else:
            menu_item.click_input()

    def select_tab_control(self, tab_control_value, parent_tab_auto_id=None):
        """ 
        Description:
        Select Tab using given EXACT value
        Example:
        | Select Tab Control | Fare 1 |
        """
        logger.info("Selecting '{}' tab".format(tab_control_value))
        syex = SyexUiaLibrary()
        syex.create_power_express_handle()
        try:
            syex.click_tabitem_control(tab_control_value, parent_tab_auto_id)      
        except LookupError:
            self.builtin.fail("Tab '{}' is not visible".format(tab_control_value))

    def select_panel(self, panel_name):
        """ 
        Selects Panel item
        
        Example:
        | Select Panel | Cust Refs |
        """
        logger.info("Selecting panel : {}".format(panel_name))
        syex = SyexUiaLibrary()
        syex.create_power_express_handle()
        syex.click_tabitem_control(panel_name)
        # time.sleep(2)

    def is_panel_selected(self, panel_name):
        try:
            syexuia = uialibrary.WindowControl(AutomationId='frmMain')
            return uialibrary.TabItemControl(Name=panel_name, searchFromControl=syexuia).CurrentIsSelected()
        except LookupError:
            self.builtin.fail('Panel name "{}" not found'.format(panel_name))

    def is_tab_visible(self, tab_value):
        tab_control_texts = self._create_app_instance().TabControl.Texts()
        return True if tab_value in tab_control_texts else False

    def get_lines_using_regexp(self, string, regex_pattern):
        """ 
        Description:
        Returns lines matching regex or pattern

        Usage:
        | Get Lines Using Regexp | ${pnr_details} | (QU|QA)/QE/PARWL24CA/70C4 |
        """
        regex_pattern = re.compile(regex_pattern)
        logger.info("Searching pattern: '{}'".format(regex_pattern.pattern))
        line_match = [line.strip() for line in string.splitlines() if re.search(regex_pattern, line)]
        return '\n'.join(line_match)

    def flatten_string(self, string_to_flatten, remove_space=False):
        """ 
        Description:
        Converts multiple lines into one line 

        Usage:
        | ${sample} | Flatten String | ${string_to_flatten} |
        
        """
        string_to_flatten = re.sub(r"([\n ])\1*", r"\1", string_to_flatten)
        if remove_space == "True":
            flattened_string = [line.strip() for line in string_to_flatten.splitlines()]
        else:
            flattened_string = [line.rstrip() for line in string_to_flatten.splitlines()]
        print "".join(flattened_string)
        return "".join(flattened_string)

    def verify_list_values_are_identical(self, list_):
        """ 
        Description:
        Returns Boolean value if list contains similar value

        Usage:
        | ${is_similar_value} | Verify List Values Are Identical  | ${list_sample} |
        
        """
        if list_: 
            return len(set(list_)) <= 1
        else:
            self.builtin.fail("List does not contain any value")

    def get_summary_texts(self):
        """ 
        Description:
        Returns summary text in list data type

        Usage:
        | ${summary_texts} = | Get Summary Texts  |
        
        """        
        # appwindow = self._create_power_express_backend_app()
        uia = SyexUiaLibrary().connect_to_express_via_uia()        
        summary_field = uia.child_window(auto_id="lvwSummary", control_type="List")
        summary_texts = [item for sublist in summary_field.texts() for item in sublist]
        return summary_texts

    def get_document_items(self):
        """ 
        Description:
        Retrieve all information from Unused Document tab

        Usage:
        | ${actual_document_items} = | Get Document Items |
        """    
        treeview_dlg = self._create_app_instance().TreeView
        return treeview_dlg.print_items().strip()

    def click_document_item(self, document_item, double_click=False):
        """ 
        Description:
        Clicks item from Unused Document

        Usage:
        | Click Document Item | PA S 56TGHY  079 1212121212                  USD3434     20180606  FULL | double_click=False |
        """
        treeview_dlg = self._create_app_instance().TreeView
        try:
            tree_item = treeview_dlg.GetItem([u''.join(document_item)])
            logger.info("selecting '{}'".format(document_item))
            if double_click == "True":
                tree_item.ClickInput(double=True)
            else:
                tree_item.ClickInput()
        except (IndexError):
            self.builtin.fail("Document item '{}' is not found and cannot be selected".format(document_item))

    def get_service_option_items(self):
        """ 
        Description:
        Gets all values of service option delimited by Tab and every item is splitted into lines
        Returns value in list

        Usage:
        | ${service_option_values}= | Get Service Option Items |

        =>
        High Fare Calculation   WPNC‡XR¦1S  High Fare
        Low Fare International Calculation  WPNIN/T2¦1S Low Fare
        """
        list_items = self._create_app_instance().list_view.texts()
        so_list = list_items[1:]
        return [so_list[index: index+2] for index, _ in enumerate(so_list)\
                     if index % 3 == 0]

    def get_fare_restiction_selected_radio_button(self, translate_to_english=True):
        """ 
        Description:
        Gets the Fare Rectrion name currently selected radio button
        Add argument True if translation to English is needed otherwise will return current language

        Usage:
        | ${currently_selected_fare_restriction} | Get Fare Restiction Selected Radio Button |
        | ${currently_selected_fare_restriction} | Get Fare Restiction Selected Radio Button | True |

        """        
        uia = SyexUiaLibrary().connect_to_express_via_uia()        
        restriction_value = ""
        for restriction in uia.child_window(auto_id="Panel1", control_type="Pane", found_index=0).children_texts():
            radio_button_state = uia.window(title=restriction,  control_type="RadioButton").is_selected()
            if radio_button_state == 1:
                restriction_value += restriction
                break
        return self._translate_fare_restriction(restriction_value, translate_to_english)    

    def _translate_fare_restriction(self, restriction_value, translate_to_english):
        new_restriction_value = ""
        if translate_to_english == "True":
            if restriction_value == "Modifiable":
                new_restriction_value += "Fully Flexible"
            elif restriction_value == "Partiellement modifiable":
                new_restriction_value += "Semi Flexible"
            elif restriction_value == "Non modifiable":
                new_restriction_value += "Non Flexible"
        else:
            new_restriction_value += restriction_value
        return new_restriction_value

    def truncate(self, value, precision=2):
        """
        Truncate the value without rounding up. By default precision used is 2.

        Example:
        ${truncated_value} = | Truncate | 15.0909 | precision=2

        Output:
        ${truncated_value} = 15.09
        """
        try:
            s = '{}'.format(value)
            if 'e' in s or 'E' in s:
                return '{0:.{1}value}'.format(value, precision)
            i, p, d = s.partition('.')
            return '.'.join([i, (d+'0'*precision)[:precision]])
        except Exception as error:
            raise ValueError(repr(error))

    def get_country_code_based_on_currency(self, currency):
        return getattr(CountryConstant, currency.lower())

    def select_gds_value(self, gds):
        logger.info("Selecting '{}' from gds toolstrip".format(gds))
        syex = SyexUiaLibrary().create_power_express_handle()
        toolstrip_button = uialibrary\
            .ButtonControl(searchFromControl=syex, RegexName='ToolStripSplitButton1|Bandeau de partage du bandeau outil1')
        toolstrip_button.SetFocus()
        toolstrip_button.Click(-10, -10)
        toolstrip_button.Click(-5, -10)
        toolstrip_button.SendKeys('{Tab}')
        gds_mapping = {'apollo':'p','amadeus':'a', 'galileo':'g', 'sabre':'s'}
        selected_gds = gds_mapping.get(gds)
        toolstrip_button.SendKeys(selected_gds)

    def get_data_from_clipboard(self):
        time.sleep(2)
        try:
            win32clipboard.OpenClipboard()
            try:
                if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_UNICODETEXT):
                    data = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
                    return data
                else:
                    return "Empty"
            finally:
                win32clipboard.CloseClipboard()
        except ValueError, err:
            raise AssertionError(err)
        
    def clear_data_from_clipboard(self):
        try:
            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.CloseClipboard()
        except ValueError, err:
            raise AssertionError(err)

    def convert_list_to_lines(self, list_):
        return "\n".join(list_)

    def remove_empty_value_from_list(self, list_):
        return filter(None, list_)

    def remove_leading_and_ending_spaces(self, item):
        return item.strip()

    def remove_all_spaces(self, item):
        return item.replace(" ", "")

    def get_latest_command_response(self, gds_data, gds=None):      
        if gds == 'sabre':
            start_index = 0
            for i, v in reversed(list(enumerate(gds_data))):
                if v.startswith(">"):
                    start_index = i
                    break
            latest_response =  gds_data[start_index:]
            print latest_response
            return self.convert_list_to_lines(latest_response)
        else:
            if gds == 'galileo':
                index_list = [i for i, line in enumerate(gds_data) if line.startswith(">") or line.startswith(')>')]
                print index_list
            else:
                reg_exp = re.compile(r">.*PAGE.*\d\/.*\d")
                index_list = [i for i, line in enumerate(gds_data) if line.startswith(">") and not line.startswith(">>")\
                         and not re.match(reg_exp, line)]
            mdr_pattern = re.compile(r"\)>mdr|>mdr|\)>|>")
            if len(index_list) > 1:
                last_two_index_list = index_list[-2:]
                latest_response = filter(lambda k: not re.match(mdr_pattern, k), gds_data[last_two_index_list[0]:])
                return self.convert_list_to_lines(latest_response)
            else:
                return self.convert_list_to_lines(filter(lambda k: not re.match(mdr_pattern, k), gds_data))

    def are_there_duplicate_remarks(self, pnr_details, gds):
        # TODO
        # implement check for other gds
        """ 
        Notes:
        Amadeus - Remarks that starts with 'RM' will be checked

        """
        remarks_list = []
        pnr_log = Counter(pnr_log.strip().lower() for pnr_log in pnr_details.split('\n') if pnr_log.strip())
        for line in pnr_log:
            if pnr_log[line] > 1:
                if gds.lower() == 'amadeus' and line != 'rir *' and line.startswith("rm"):
                    print 'duplicate remarks found: ' + line
                    remarks_list.append(line)
        return True if len(remarks_list) > 0 else False

    def sort_pnr_details(self, pnr_details):
        """ 
        Description:
        Sorts PNR Details and removes duplicate line

        Example:
        | ${sorted_pnr_details} = | Sort Pnr Details | ${pnr_details}
        """
        remark_list = [line for line in pnr_details.splitlines()]
        # for index, remark in enumerate(remark_list):
        #     if re.match(r'\s{1,4}\d', remark):
        #         remark_list[index] = remark.lstrip()
        # seen = set()
        # seen_add = seen.add
        # sorted_pnr_details = [remark for remark in remark_list \
                            # if not (remark in seen or seen_add(remark))]
                            # or re.match(r'[ \t\t]', remark)]
        return "\n".join(remark_list)

    def verify_no_duplicate_remarks_exists(self, pnr_details, gds=None):
        """ 
        Description:
        Verifies if two consecutive lines are duplicate. Displays remark if duplicate is found
        """
        actual_remarks_list = filter(None, [re.sub(r'(^\d*\.?\s?)', '', line) for line in pnr_details.splitlines()])
        duplicate_remark_list = []
        for i, value in enumerate(actual_remarks_list):
            current_elem = value
            next_elem = actual_remarks_list[(i + 1) % len(actual_remarks_list)]
            if current_elem == next_elem:
                duplicate_remark_list.append(current_elem)

        duplicate_remark_list_length = len(duplicate_remark_list)
        if duplicate_remark_list_length > 1:
            for elem in duplicate_remark_list:
                print elem
        self.builtin.should_not_be_true(duplicate_remark_list_length > 1)

    def round_apac(self, number, country):
        if '.' in str(number):
            number_split = str(number).split('.')
            whole_num = number_split[0]
            decimal = number_split[1]
            if country.upper() == 'HK' or country.upper() == 'IN':
                if int(decimal[0]) > 4:
                    return int(whole_num) + 1
                else:
                    return int(whole_num)
            elif country.upper() == 'SG':
                if len(decimal) < 3:
                    return self.convert_to_float(number)
                else:
                    if decimal[2] == '5':
                        return whole_num + '.' + str(int(decimal[:2]) + 1)
                    else:
                        return self.convert_to_float(number)
                        # return round(number, 2)
        else:
            if country.upper() == 'HK' or country.upper() == 'IN':
                return int(number)
            elif country.upper() == 'SG':
                return self.convert_to_float(number)

    def select_gds_screen_tab(self):
        try:
            app = self._create_app_instance()
            app["CRS Screen"].TabControl.Select(0)
        except Exception:
            pass

    def select_cell_in_data_grid_table(self, control_id, selected_index):
        """ 
        Description:
        Simulates the User Action Click on Grid Tables to select single cell. 

        (Note that the minimum selected_index value is 0)

        """
        select_row = SyexUiaLibrary().get_table_control(control_id).CustomControl(Name=selected_index)
        select_row.SetFocus()
        select_row.Click()

    def select_multiple_cells_in_data_grid_table(self, control_id, index_list):
        """ 
        Description:
        Simulates the User Action Ctrl+ Click on Grid Tables to select multiple cells.

        index_list is the list of indices to select to.  (Note that the minimum index_list value is 0)

        """
        table_control = SyexUiaLibrary().get_table_control(control_id)
        table_control.SetFocus()
        
        for i, index in enumerate(index_list):
            table_control.CustomControl(Name=str(index)).Click()
            if i == 0:
                autoit.send("{CTRLDOWN}")
        autoit.send("{CTRLUP}")

    def get_cell_value_in_data_grid_table(self, control_id, selected_index):
        """ 
        Description:
        Retrieve the value of the selected Cell in a Grid Table

        (Note that the minimum selected_index value is 0)

        """
        select_row = SyexUiaLibrary().get_table_control(control_id).CustomControl(Name=selected_index)
        return select_row.AccessibleCurrentValue()

    def click_cell_in_data_grid_pane(self, control_id, selected_row, selected_column):
        """ 
        Description:
        Simulates the User Action Click on Grid Panes to click a certain cell.

        (Note that the minimum selected_row/selected_column value is 0)

        """
        select_row = SyexUiaLibrary().get_pane_control(control_id).CustomControl(Name=selected_row)
        select_row.GetChildren()[int(selected_column)].SetFocus()
        select_row.GetChildren()[int(selected_column)].Click()

    def get_cell_value_in_data_grid_pane(self, control_id, selected_row, selected_column):
        """ 
        Description:
        Get the value of the selected Cell in a Grid Pane

        (Note that the minimum selected_row/selected_column value is 0)

        """
        select_row = SyexUiaLibrary().get_pane_control(control_id).CustomControl(Name=selected_row)
        return select_row.GetChildren()[int(selected_column)].AccessibleCurrentValue()
        
    def get_all_cell_values_in_data_grid_table(self, control_id):
        """ 
        Description:
        Retrieve the values of all cell values in a Grid Table

        """
        select_table = SyexUiaLibrary().get_table_control(control_id)
        return [x.AccessibleCurrentValue().replace('(null)', '') for x in select_table.GetChildren() if x.AccessibleCurrentValue() != "0"]

    def get_all_cell_values_in_data_grid_pane(self, control_id, column_index=2):
        """ 
        Description:
        Retrieve the values of all cell values in a Grid Pane 

        Note:
        This is specific to Grid Panes in Remarks Tab (Other Services where the texts are located at third column)

        """
        select_pane = SyexUiaLibrary().get_pane_control(control_id)
        select_pane_list = [x.GetChildren()[int(column_index)].AccessibleCurrentValue() for x in select_pane.GetChildren()]
        return filter(None, select_pane_list)

    def _get_name(self, control_id, number_of_records=5):
        # Need to be revisit the logic
        list = []
        for x in range(int(number_of_records)):
            list.append(SyexUiaLibrary().get_table_control(control_id).CustomControl(Name='Name Row ' + str(x)).AccessibleCurrentValue())    
        return list

    def delete_record_in_table_grid(self, control_field, record_to_delete, actual_number_of_records):
        # Need to be revisit the logic
        current_records = self._get_name(control_field, actual_number_of_records)
        autoit.send("{DOWN}" * int(current_records.index(record_to_delete)))
        for x in SyexUiaLibrary().get_table_control(control_field).GetChildren():
          if x.ControlTypeName == "CustomControl":
              if x.Name == "Row " + str(current_records.index(record_to_delete)):
                  for y in x.GetChildren():
                      if y.AccessibleCurrentValue() == record_to_delete:
                          y.GetPreviousSiblingControl().Click()
    
    def get_segments_from_list_control(self):
        """
        Description: Get All Segments from the Grid
        | ${segments} | Get Segments In Data Grid  |
        
        Output:
        [u'2 PR 510 Y 25FEB 1 SINMNL HK1 0020 0355']
        """     
        segments = [segment.Name for segment in uialibrary.ListControl(AutomationId='SegmentListView').GetChildren() if segment.Name != 'Header Control' and segment.Name != 'Vertical' and segment.Name != 'Horizontal' and segment.Name != '']
        return segments

    def click_amend_eo_button(self, eo_number):
        """
        Search for specified EO number and Click the Amend button
        | Click Amend EO | 1808100081  |
        
        """     
        SyexUiaLibrary().create_power_express_handle()
        try:
            row = self.get_row_object_from_datagrid(eo_number, 'EoGrid')
            row[1].SetFocus()
            row[1].Click()       
        except LookupError:
            self.builtin.fail("No EO Record.")
        # finally:
        #     ImageHorizonLibrary().take_a_screenshot()

    def delete_associated_charges_record(self, control_field, record):
        """
        Delete existing record from the grid
        | Delete Record In Associated Charges | VAT  |
        
        """     
        SyexUiaLibrary().create_power_express_handle()
        control_field = 'AssociatedCharges_AssociatedInfoGrid'
        row = self.get_row_object_from_datagrid(record, control_field)
        row[1].Click()

    def delete_vendor_grid_record(self, control_field, detail_name):
        """
        Delete existing record from the grid
        | Delete Record In Vendor Tab |
        
        """     
        SyexUiaLibrary().create_power_express_handle()
        control_field = 'ContactTypeDataGridView'
        row = self.get_row_object_from_datagrid(detail_name, control_field)
        row[1].Click()

    def delete_train_grid_record(self, control_field, train_name):
        """
        Delete existing record from the grid
        | Delete Record In Request Tab |

        """
        SyexUiaLibrary().create_power_express_handle()
        row = self.get_row_object_from_datagrid(train_name, control_field)
        row[1].Click()
        # uialibrary.ButtonControl(Name="Yes")

    def _get_eo_grid_handle(self):
        syex = SyexUiaLibrary().create_power_express_handle()
        eo_datagrid = uialibrary.TableControl(AutomationId='FromEoDataGridView', searchFromControl=syex)
        return eo_datagrid

    def _get_io_grid_handle(self):
        syex = SyexUiaLibrary().create_power_express_handle()
        io_datagrid = uialibrary.TableControl(AutomationId='FromIoDataGridView', searchFromControl=syex)
        return io_datagrid

    def get_all_remarks_from_eo_grid(self):
        eo_datagrid_children = self._get_eo_grid_handle().GetChildren()
        return [x.AccessibleCurrentValue() for x in eo_datagrid_children
                if x.ControlType == uialibrary.ControlType.CustomControl]

    def get_all_remarks_from_io_grid(self):
        io_datagrid_children = self._get_io_grid_handle().GetChildren()
        return [x.AccessibleCurrentValue() for x in io_datagrid_children
                if x.ControlType == uialibrary.ControlType.CustomControl]

    def select_remarks_from_eo_grid(self, *remarks):
        eo_remarks = self.get_all_remarks_from_eo_grid()
        eo_datagrid = self._get_eo_grid_handle()
        for remark in remarks:
            try:
                remark_index = eo_remarks.index(remark)
            except ValueError:
                self.builtin.fail("Remark '{}' not found on grid".format(remark))
            logger.info("Selecting '{}' remark from eo grid".format(remark))
            eo_datagrid.SetFocus()
            eo_datagrid.SendKeys("{CTRL}{HOME}")
            if remark_index > 0:
                eo_datagrid.SendKeys("{DOWN}" * int(remark_index))
            uialibrary.ButtonControl(AutomationId='EoAddButton').Click()
           
    def select_remarks_from_io_grid(self, *remarks):
        io_remarks = self.get_all_remarks_from_io_grid()
        io_datagrid = self._get_io_grid_handle()
        for remark in remarks:                    
            try:
                remark_index = io_remarks.index(remark)
            except ValueError:
                self.builtin.fail("Remark '{}' not found on grid".format(remark))
            logger.info("Selecting '{}' remark from io grid".format(remark))
            io_datagrid.SetFocus()
            io_datagrid.SendKeys("{CTRL}{HOME}")
            if remark_index > 0:
                io_datagrid.SendKeys("{DOWN}" * int(remark_index))
            uialibrary.ButtonControl(AutomationId='IoAddButton').Click()
            
    def click_send_email_eo(self, eo_number):
        """
        Search for specified EO number and Click the Amend button
        | Click Send Email EO | 1808100081  |
        
        """     
        SyexUiaLibrary().create_power_express_handle()
        row = self.get_row_object_from_datagrid(eo_number, 'EoGrid')
        row[2].Click()

    def convert_segment_number_to_gds_format(self):
        """
        Get All Selected Segments From Select Air Segments Grid

        Example Selected Segments
           Selected (X)     | '2 PR 510 Y 07MAR 4 SINMNL HK1 0020 0355'
           Not Selected (0) | '3 PR 300 Y 10MAR 7 MNLHKG HK1 0755 1005'
           Selected (X)     | '4 PR 313 Y 15MAR 5 HKGMNL HK1 0750 1000'
        Output:
            segment_number_short = 2,4
            segment_number_long = 0204

        Example Selected Segments
           Selected (X)     | '2 PR 510 Y 07MAR 4 SINMNL HK1 0020 0355'
           Selected (X)     | '3 PR 300 Y 10MAR 7 MNLHKG HK1 0755 1005'
           Selected (X)     | '4 PR 313 Y 15MAR 5 HKGMNL HK1 0750 1000'
        Output:
            segment_number_short = 2-4
            segment_number_long = 020304

        | ${segment_number_short} | ${segment_number_long} | Convert Segment Number To GDS Format  |
        
        """

        selected_segments = self._get_segment_selected()
        segment_numbers = [number.split(' ')[0] for number in selected_segments]
        gds_segment_number_long = ['0' + number for number in segment_numbers]
        gds_segment_number_long = ''.join(gds_segment_number_long)

        if len(selected_segments) == 1:
            gds_segment_number_short = ''.join(segment_numbers)
        else:
            is_sequential = self._is_segment_number_sequential(segment_numbers)

            if is_sequential == True:
                gds_segment_number_short = segment_numbers[0] + "-" + segment_numbers[-1]
            else:
                gds_segment_number_short = ",".join(segment_numbers)

        return gds_segment_number_short, gds_segment_number_long

    def _is_segment_number_sequential(self, _segment_numbers):
        is_sequential = False
        for i, number in enumerate(_segment_numbers, 1):
            if int(number)-1 == i:
                is_sequential = True
            else:
                is_sequential = False
                break   
        return is_sequential

    def _get_segment_selected(self):
        segment_list = SyexCustomLibrary().get_segments_from_list_control()
        segments_selected_list = []
        for segment in segment_list:
            if segment != 'Header Control':
                if SyexUiaLibrary().get_checkbox_state_from_list_control(segment):
                    segments_selected_list.append(segment)
        return segments_selected_list

    def get_equivalent_city_name(self, excel_file, city_code):
        wb = load_workbook(excel_file)
        ws = wb.worksheets[0]
        city_codes_list = [cell.value for cell in [col for col in ws['A']]]
        city_names_list = [cell.value for cell in [col for col in ws['B']]]
        city_dict = dict(zip(city_codes_list, city_names_list))
        return city_dict.get(city_code)

    def get_remarks_set_from_pnr_details(self, pnr_details, start_marker, end_marker):
        # return re.findall('('+re.escape(start_marker)+'(.*?)'+ re.escape(end_marker) + ')', 
        #     pnr_details)
        logger.info(start_marker)
        logger.info(end_marker)
        first_marker = pnr_details.index(start_marker) + len(start_marker)
        last_marker = pnr_details.index(end_marker, first_marker)
        remarks_set = start_marker +  pnr_details[first_marker:last_marker] + end_marker
        remarks_set_list = filter(None, [re.sub(r"(^\s+?(\d+\.?)\s+)", '', line) \
                                for line in remarks_set.splitlines()])
        return "\n".join(remarks_set_list)

    def _get_mco_mpd_handle(self):
        mco_mpd_grid = SyexUiaLibrary().get_pane_control('RemarksDataGridView')
        return mco_mpd_grid

    def get_all_mco_mpd_remarks_from_remarks_grid(self):
        mco_mpd_children = self._get_mco_mpd_handle().GetChildren()
        children = [child.GetLastChildControl() for child in mco_mpd_children]
        remarks = [child.AccessibleCurrentValue() for child in children]
        return remarks

    def click_list_item(self, item, double_click=True):
        """
        Description: Accepts argument as regex or literal value 
        | Click List Item | 2                          | double_click=True |
        | Click List Item | 2. PR 22FEB2019 SINMNL HK1 |                   |
        """        
        if unicode(item).isnumeric():

            segment_list = SyexCustomLibrary().get_segments_from_list_control()
            segment_numbers = [number.split(' ')[0] for number in segment_list]
            segment_index = segment_numbers.index(item)

            if str(double_click) == "True":
                uialibrary.ListItemControl(Name=segment_list[segment_index]).DoubleClick()
            else:
                uialibrary.ListItemControl(Name=segment_list[segment_index]).Click()
        else:
            if str(double_click) == "True":
                uialibrary.ListItemControl(RegexName='{}'.format(item)).DoubleClick()
            else:
                uialibrary.ListItemControl(RegexName='{}'.format(item)).Click()

    def is_error_icon_visible(self, parent_auto_id):
        """
        Description: Use the parent id of error icon
        | ${is_ticketing_date_error } | Is Error Icon Visible | is_error_icon_visible |

        Returns boolean
        """            
        return bool(re.search(r'\d', uialibrary.PaneControl(AutomationId=parent_auto_id)\
                        .GetFirstChildControl().AutomationId))

    def delete_itinerary_remarks_in_air_restrictions(self, control_id, remark_value, remark_instance=None):
        """
        Description: 
        Deletes Remarks present in the Itinerary Remarks Grid In AirFare and Alternate AirFare Restriction Panel
        If remarks_instance is None, Deletes All the Remarks with Remark Name passed in Arguments.
        | Delete Itinerary Remarks In Air Restrictions | Control | Remark |                      |
        | Delete Itinerary Remarks In Air Restrictions | Control | Remark | 1                    |
        """        
        control_id = SyexUiaLibrary()._convert_control_id_to_uia_format(control_id)
        itin_remarks_table = uialibrary.PaneControl(AutomationId=control_id)
        itin_remarks_table.SetFocus()
        
        parent_match_list = []
        for item in itin_remarks_table.GetChildren():
            for row in item.GetChildren():
                if row.AccessibleCurrentValue() == remark_value:
                    parent_match_list.append(row)
        
        if remark_instance is None:
            for each_parent in parent_match_list:
                try:
                    next_sibling = each_parent.GetNextSiblingControl()
                except UnboundLocalError:
                    self.builtin.fail("Data :{} is not present in grid".format(remark_value))
                next_sibling.Click()
        else:
            parent_instance = parent_match_list[int(remark_instance) - 1]
            try:
                next_sibling = parent_instance.GetNextSiblingControl()
            except UnboundLocalError:
                self.builtin.fail("Data :{} is not present in grid".format(remark_value))
            next_sibling.Click() 
  
    def click_button_in_eo_grid(self, eo_number, country_code, eo_action):
        """
        Search for specified EO number and Click the Raise Cheque button
        | Click EO Grid Button | 1808100081  | Amend
        | Click EO Grid Button | 1808100081  | PDF
        | Click EO Grid Button | 1808100081  | Email
        | Click EO Grid Button | 1808100081  | Raise Cheque
        | Click EO Grid Button | 1808100081  | Cancel
        
        """  
        if country_code.upper() == "IN":
            button_menu_dict = {
                "amend" : 1,
                "email" : 2, 
                "raise cheque": 3,
                "cancel": 4
            }
        else:
            button_menu_dict = {
                "amend": 1, 
                "pdf": 2, 
                "email": 3, 
                "raise cheque": 4, 
                "cancel": 5
                }

        eo_action = eo_action.lower()
        button_index = button_menu_dict.get(eo_action,6)
        self.get_row_in_eo_grid(eo_number)[button_index].Click()

    def _get_datagrid_children(self, grid_id):
        grid_parent = SyexUiaLibrary().get_table_control(grid_id).GetChildren()
        return [x for x in grid_parent if x.ControlType == uialibrary.ControlType.CustomControl][1:]

    def get_row_object_from_datagrid(self, item=None, grid_id=None):
        if item is not None:
            for selected_row in self._get_datagrid_children(grid_id):
                for column_name in selected_row.GetChildren():
                    if column_name.AccessibleCurrentValue() == item:
                        row_object = selected_row
            return row_object.GetChildren()
        else:
            return [x.AccessibleCurrentValue() for row in self._get_datagrid_children(grid_id) \
                for x in row.GetChildren() if x.ControlType == uialibrary.ControlType.CustomControl]

    def get_all_values_from_datagrid(self, grid_id):
        """
        Get all values and store it in list type
        """
        return filter(lambda l: 'System.Drawing.Bitmap' not in l, self.get_row_object_from_datagrid(grid_id=grid_id))

    def get_row_in_eo_grid(self, eo_number):
        return self.get_row_object_from_datagrid(eo_number, 'EoGrid')

    def get_row_values_from_datagrid(self, item, datagrid_id):
        datagrid_row = self.get_row_object_from_datagrid(item, datagrid_id)
        return filter(lambda x: 'System.Drawing.Bitmap' not in x, [v.AccessibleCurrentValue().\
                replace('(null)', '') for v in datagrid_row])

    def get_row_values_in_eo_grid(self, eo_number):
        return self.get_row_values_from_datagrid(eo_number, 'EoGrid')

    def get_row_fields_from_custom_fields(self):
        row_values = []
        for each_child in self._get_datagrid_children("CustomFieldGrid"):
            row_values.append(each_child.AccessibleCurrentValue())
        return row_values

    def set_row_object_in_datagrid(self, row_item, row_value, grid_auto_id):
        """
        Sets value of each datagrid in custom fields
        Sample: Set Row Object In Datagrid  <Row name from UI> <Value to be Set> <Control Id name>
        """
        SyexUiaLibrary().create_power_express_handle()
        actual_row_object = self.get_row_object_from_datagrid(row_item, grid_auto_id)
        for v in actual_row_object:
            if v.AccessibleCurrentValue() == row_item:
                v.Click()
                autoit.send('{TAB}')
                autoit.send(row_value)

    def round_to_nearest_dollar(self, amount, country, round_type=None):
        """
        If round type is not specified, it will be rounded according to standard
        If country is SG, round value will return float else returns integer
        | ${round_down_value}| round to nearest dollar | 123.34 | SG | down |
        | ${round_up_value}  | round to nearest dollar | 123.34 | HK | up   |
        | ${round_value}     | round to nearest dollar | 123.34 | IN |      |

        >>
        ${round_down_value} = 2334.00
        ${round_up_value} = 2335.00
        ${round_value} = 2334.00
        """          
        round_type_dict= {'up':'ROUND_UP', 'down':'ROUND_DOWN'}
        amount_in_decimal = Decimal(amount)
        round_type = round_type_dict.get(str(round_type).lower(), 'ROUND_HALF_UP')
        round_value =  amount_in_decimal.quantize(Decimal('1'), rounding=round_type).quantize(Decimal('0.01'))
        return round_value if country.upper() == 'SG' else int(round_value)

# if __name__ == '__main__':
#     syex = SyexCustomLibrary()
