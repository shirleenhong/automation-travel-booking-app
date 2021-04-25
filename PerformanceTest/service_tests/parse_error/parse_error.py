import os
import csv
import datetime
from openpyxl import Workbook
import re
import ast
import sys

# logfile = "apac_services_1000.log"

def parse_error(logfile):
	wb = Workbook()
	ws = wb.worksheets[0]
	ws['A1'] = 'Request Type'
	ws['B1'] = 'Name'
	ws['C1'] = 'Exception'

	with open(logfile) as f:
		contents = f.read().split("\n")
		row_number = 1
		for line in contents:
			if '"result":"ERR"' in line:
				line = re.sub(r".*stdout:\s", "", line)
				line_dict = ast.literal_eval(line)
				row_number += 1			
				ws['A' + str(row_number)] = line_dict.get("request_type")
				ws['B' + str(row_number)] = line_dict.get("name")
				ws['C' + str(row_number)] = line_dict.get("exception")


	date_stamp = '{:%Y%m%d_%H%M%S}'.format(datetime.datetime.now())
	wb.save(filename=logfile.split(".")[0] + "_" + date_stamp +'.xlsx')

parse_error(logfile)