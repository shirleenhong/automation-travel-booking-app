# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color
from openpyxl.worksheet import dimensions
from collections import defaultdict
import glob
import datetime
import getpass
import re
import sys
_global = sys.modules[__name__]

class SyexPerfTestLibrary:

    ROBOT_LIBRARY_SCOPE = 'Test Suite'

    def generate_time_stamp(self):
        return '{:%Y%m%d%H%M%S}'.format(datetime.datetime.now())

    def rename_file(self, from_file, to_file):
        os.rename(from_file, to_file)

    def get_perf_logs(self, environment, version, log_file=''):
        _global.environment = environment
        if log_file == '':
            local_app_data = os.getenv('LOCALAPPDATA') + "\\Carlson Wagonlit Travel\\Power Express\\"+ version +"\\"
            file_name = "{0}.SyExPerfLog_{1}.log".format(getpass.getuser(), datetime.date.today().strftime('%Y%m%d'))
            log_file = "".join(glob.glob(local_app_data + file_name))
        regex_classic_comm = re.compile('Performance="(\d*)"\s+Units="msec" ClassName="Gds-AmadeusCommunication" MethodName="(.*)"')
        regex_sellco_comm = re.compile('Performance="(\d*)"\s+Units="msec" ClassName="Gds-AmadeusSecoCommunication" MethodName="(.*)"')
        regex_bfm = re.compile('Performance="(\d*)"\s+Units="msec" ClassName="BusinessFunctionMetric" MethodName="(.*)"')

        with open(log_file) as f:
            if environment == 'sellco':
                matches_amadeus = filter(None, [regex_sellco_comm.findall(line) + regex_bfm.findall(line) for line in f])
            else:
                matches_amadeus = filter(None, [regex_classic_comm.findall(line) + regex_bfm.findall(line) for line in f])
        return matches_amadeus


    def save_perf_logs_to_excel(self, number_of_execution, *args):
        username = getpass.getuser()
        result_file = "{0}_{1}_perf_log_{2}.xlsx".format(username, environment, self.generate_time_stamp())
        wb = Workbook()
        for i in range(int(number_of_execution)):
            ws = wb.create_sheet("TestRun" + str(i + 1))
            ws['A1'] = 'Response'
            ws['B1'] = 'Method Name'
            ws['G1'] = 'Response'
            ws['H1'] = 'Method Name'
            for index, rs_service in enumerate(args[0][i]):
                if "BusinessFunctionMetric" not in rs_service[0][1]:
                    ws['A' + str(index + 2)] = int(rs_service[0][0])
                    ws['B' + str(index + 2)] = "".join(rs_service[0][1])
                else:
                    ws['G' + str(index + 2)] = int(rs_service[0][0])
                    ws['H' + str(index + 2)] = "".join(rs_service[0][1])
            # ws.auto_filter.ref = 'A1:B1'
            # ws.auto_filter.add_filter_column(1, ['SendCryptic'])
        ws = wb.worksheets[0]
        ws['B1'] = 'Total GDS Communication RS'
        ws['C1'] = 'Average GDS Communication RS'
        ws['D1'] = 'Max RS - GDS Communication Method'
        ws['E1'] = 'Max RS'
        ws['F1'] = 'Total Business Function Metric RS'
        ws['G1'] = 'Average Business Function Metric RS'
        ws['H1'] = 'Max RS - Business Function Metric Method'
        ws['I1'] = 'Max RS'
        self._format_header(ws, 'B1:I1')
        for i in range(1, int(number_of_execution) + 1):
            ws['A' + str(int(i + 1))] = "TestRun"+ str(i)
            ws['B' + str(int(i + 1))] = "=SUM(TestRun"+ str(i) +"!$A:$A)"
            ws['C' + str(int(i + 1))] = "=AVERAGE(TestRun"+ str(i) +"!$A:$A)"
            ws['D' + str(int(i + 1))] = "=VLOOKUP(MAX(TestRun"+ str(i) +"!$A:$A),TestRun"+ str(i) +"!$A:$B,2,FALSE)"
            ws['E' + str(int(i + 1))] = "=MAX(TestRun"+ str(i) +"!$A:$A)"
            ws['F' + str(int(i + 1))] = "=SUM(TestRun"+ str(i) +"!$G:$G)"
            ws['G' + str(int(i + 1))] = "=AVERAGE(TestRun"+ str(i) +"!$G:$G)"
            ws['H' + str(int(i + 1))] = "=VLOOKUP(MAX(TestRun"+ str(i) +"!$G:$G),TestRun"+ str(i) +"!$G:$H,2,FALSE)"
            ws['I' + str(int(i + 1))] = "=MAX(TestRun"+ str(i) +"!$G:$G)"
        last_row = ws.max_row
        summary_row = last_row + 1
        ws['A{}'.format(summary_row)] = 'Total'
        ws['B{}'.format(summary_row)] = "=SUM(B2:B{})".format(last_row)
        ws['C{}'.format(summary_row)] = "=SUM(C2:C{})".format(last_row)
        ws['F{}'.format(summary_row)] = "=SUM(F2:F{})".format(last_row)
        ws['G{}'.format(summary_row)] = "=SUM(G2:G{})".format(last_row)
        self._format_header(ws, 'A{0}:G{0}'.format(summary_row))
        wb.save(filename=result_file)
        self.parse_excel_file(result_file)


    def parse_excel_file(self, result_file):
        wb = load_workbook(result_file)
        ws = wb.worksheets[0]
        max_r = ws.max_row + 2
        for i, value in enumerate(['Method Name', 'Average RS', 'Minimum RS', 'Maximum RS']):
            ws.cell(column=i+1, row=max_r, value=value)
        self._format_header(ws, 'A{0}:D{0}'.format(ws.max_row))

        cryptic_dict = defaultdict(list)
        non_cryptic_dict = defaultdict(list)
        bfm_dict = defaultdict(list)
        for sheet in wb.worksheets[1:]:
            for i in range(2, len(sheet['B']) + 1):
                conn_rs =  sheet['A' + str(i)].value
                conn_method =  sheet['B' + str(i)].value
                bfm_rs =  sheet['G' + str(i)].value
                bfm_method =  sheet['H' + str(i)].value
                if conn_rs is not None:
                    if not conn_method.startswith("SendCryptic"):
                        non_cryptic_dict[conn_method].append(int(conn_rs))
                    else:
                        cryptic_dict[conn_method].append(int(conn_rs))
                if bfm_rs is not None:
                    bfm_dict[bfm_method].append(int(bfm_rs))

        counter_cryptic = ws.max_row + 1        
        for k,v in sorted(cryptic_dict.items()):
            ws['A'+ str(counter_cryptic)] = "".join(k)
            ws['B'+ str(counter_cryptic)] = sum(v)/float(len(v))
            ws['C'+ str(counter_cryptic)] = min(v)
            ws['D'+ str(counter_cryptic)] = max(v)
            counter_cryptic += 1
        counter_non_cryptic = ws.max_row + 2
        for k,v in sorted(non_cryptic_dict.items()):
            ws['A'+ str(counter_non_cryptic)] = "".join(k)
            ws['B'+ str(counter_non_cryptic)] = sum(v)/float(len(v))
            ws['C'+ str(counter_non_cryptic)] = min(v)
            ws['D'+ str(counter_non_cryptic)] = max(v)
            counter_non_cryptic += 1
        counter_bfm = ws.max_row + 2
        for k,v in sorted(bfm_dict.items()):
            ws['A'+ str(counter_bfm)] = "".join(k)
            ws['B'+ str(counter_bfm)] = sum(v)/float(len(v))
            ws['C'+ str(counter_bfm)] = min(v)
            ws['D'+ str(counter_bfm)] = max(v)
            counter_bfm += 1
        # dimensions.ColumnDimension(ws,auto_size=True)
        wb.save(filename=result_file)

    def _format_header(self, ws, column_range):
        for cell in ws[column_range][0]:
            cell.font = Font(size=12, bold=True)

if __name__ == '__main__':
    syex = SyexPerfTestLibrary()
    syex.open_excel()
#     # syex.get_perf_logs("sello", "18.4")
#     syex.parse_excel_file('sellco_perf_log_20180425161424.xlsx')